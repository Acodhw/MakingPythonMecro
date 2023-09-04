from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from openpyxl import Workbook

import re

import time

print("먹고 싶은 음식? : ", end="") #input What food do you want to eat
food = input()

driver = webdriver.Chrome()
driver.implicitly_wait(10) # wait for loading
driver.get('https://map.naver.com/p/') #link get
time.sleep(1)

driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[1]/div/div[1]/div/div/div/input').send_keys(food + '맛집') # send Keyword "맛집" to find restaurant
driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[1]/div/div[1]/div/div/div/input').send_keys(Keys.RETURN) # Press Enter
time.sleep(2)
iframe = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[1]/div/div[3]/div/div/div[2]/iframe') #Get iframe
driver.switch_to.frame(iframe)    
ls = driver.find_elements(By.XPATH, '/html/body/div[3]/div/div[3]/div[1]/ul/li') # Get place list

if len(ls) != 0:
    dic = {}  # make dictonary
    j = 1

    for i in ls: # loop for add dictonary value
        AllInfo=[] # KEY
        driver.switch_to.default_content() # go to default frame
        driver.switch_to.frame(iframe) # get iframe(list of place)
        
        driver2 = webdriver.Chrome() #Find rating with google because Naver Map no longer shows star ratings.
        driver2.implicitly_wait(10)
        driver2.get('https://www.google.co.kr/maps/?hl=ko&entry=ttu') #link get
        try: # catch except when can't access first xpath link
            AllInfo.append(driver.find_element(By.XPATH, '/html/body/div[3]/div/div[3]/div[1]/ul/li['+ str(j) +']/div[1]/a[1]/div/div/span[3]').text)
        except:
            AllInfo.append(driver.find_element(By.XPATH, '/html/body/div[3]/div/div[3]/div[1]/ul/li['+ str(j) +']/div[1]/a[1]/div/div/span[2]').text)
        name = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[3]/div[1]/ul/li['+ str(j) +']/div[1]/a[1]/div/div/span[1]').text      
        # name = rastaurant name
        if(name in dic): # if already get check(ex: ads)
            driver2.close()
            j+=1
            continue

        print(name)

        driver2.find_element(By.XPATH, '/html/body/div[3]/div[8]/div[3]/div[1]/div[1]/div/div[2]/form/input').send_keys(name) # Find in Google map
        driver2.find_element(By.XPATH, '/html/body/div[3]/div[8]/div[3]/div[1]/div[1]/div/div[3]/div').click()
        ReviewPoint=""
        
        try: # catch when google map can't find the place or can find rating => send Error Message
            score = driver2.find_element(By.XPATH, '/html/body/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div/div[1]/div[2]/div/div[1]/div[2]/span[1]/span[1]').text
            review = driver2.find_element(By.XPATH, '/html/body/div[3]/div[8]/div[9]/div/div/div[1]/div[2]/div/div[1]/div/div/div[2]/div/div[1]/div[2]/div/div[1]/div[2]/span[2]/span/span').text
            ReviewPoint = score + review      
        except:
            print("구글 평점과 연계되지 않습니다.")
            ReviewPoint = "구글 평점과 연계되지 않았습니다."
        driver2.close()       
        
        driver.find_element(By.XPATH, '/html/body/div[3]/div/div[3]/div[1]/ul/li['+ str(j) +']/div[1]/a[1]/div/div/span[1]').click()
        time.sleep(1)
        driver.switch_to.default_content()
        iframe2 = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[1]/div/div[3]/div/div/iframe') #Get iframe(place info)
        driver.switch_to.frame(iframe2)  

        address = ""
        menuList = ""
        reviewList = [] 
        address = driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[6]/div/div[2]/div/div/div[1]/div/a/span[1]').text # get address of place   
        menuBtn = 2     

        if(driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[5]/div/div/div/div/a[3]/span').text == "메뉴"): # Find Menu Button
            menuBtn = 3
        if(driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[5]/div/div/div/div/a[4]/span').text == "메뉴"):
            menuBtn = 4

        driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[5]/div/div/div/div/a[' + str(menuBtn) + ']/span').click() # Menu Button Click
        try: # catch when can't find menu list => find other link, catch when can't find menu list in other link => Error message get
            tmp = driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[7]/div/div[2]/div/ul').text
            menuList = tmp
        except:
            try:
                tmp = driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[7]/div[2]/div[1]/div[1]/ul').text
                menuList = tmp
            except:
                print("메뉴 정보를 불러올 수 없습니다.")
                menuList = "메뉴 정보를 불러올 수 없습니다."

        driver.switch_to.default_content() 
        driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[1]/div/div[3]/button[1]').click()
        driver.switch_to.frame(iframe) # Info window off and turn on(Reset) for avoid button link error
        driver.find_element(By.XPATH, '/html/body/div[3]/div/div[3]/div[1]/ul/li['+ str(j) +']/div[1]/a[1]/div/div/span[1]').click()
        time.sleep(1)
        driver.switch_to.default_content()
        iframe2 = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[1]/div/div[3]/div/div/iframe') #Get iframe(place info)
        driver.switch_to.frame(iframe2)  

        driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[5]/div/div/div/div/a[' +str(menuBtn+1)+ ']/span').click()
        k = 1
        try: # catch can't find review.
            ls2 = driver.find_elements(By.XPATH, '/html/body/div[3]/div/div/div/div[7]/div[3]/div[3]/div[1]/ul/li') 
            for n in ls2: # loop for getting all reviews.
                driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[7]/div[3]/div[3]/div[1]/ul/li['+ str(k) +']/div[3]/a/span').click()
                reviewList.append(driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[7]/div[3]/div[3]/div[1]/ul/li['+ str(k) +']/div[3]/a').text)
                k += 1
        except:
            print("다음 리뷰 정보를 불러올 수 없습니다.")
            reviewList.append("다음 리뷰 정보를 불러올 수 없습니다.")
        
        if(len(reviewList) == 0): # if can't get review but it is not except(ex:link is wrong, but it have anything, not text), find new link and get again.
            try:
                ls2 = driver.find_elements(By.XPATH, '/html/body/div[3]/div/div/div/div[7]/div[2]/div[3]/div[1]/ul/li') 
                for n in ls2:
                    driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[7]/div[2]/div[3]/div[1]/ul/li['+ str(k) +']/div[3]/a/span').click()
                    reviewList.append(driver.find_element(By.XPATH, '/html/body/div[3]/div/div/div/div[7]/div[2]/div[3]/div[1]/ul/li['+ str(k) +']/div[3]/a').text)
                    k += 1
            except:
                print("다음 리뷰 정보를 불러올 수 없습니다.")
                reviewList.append("다음 리뷰 정보를 불러올 수 없습니다.")

        driver.switch_to.default_content()
        driver.find_element(By.XPATH, '/html/body/div[1]/div/div[2]/div[1]/div/div[3]/button[1]').click()

        AllInfo.append(address)
        AllInfo.append(ReviewPoint)
        AllInfo.append(menuList)
        AllInfo.append(reviewList)
        dic[name] = AllInfo # add infomation

        j+=1
        print()
        time.sleep(1)

    print(dic)

    #add to Exel file
    write_wb = Workbook()
    write_ws = write_wb.create_sheet('sheet1') #creat exel file

    write_ws = write_wb.active
    alphabetValue = [ord('B')]
    write_ws['A1'] = "주제 : " + food
    write_ws['A2'] = "가게"
    write_ws['A3'] = "장르"
    write_ws['A4'] = "주소"
    write_ws['A5'] = "구글평점(인원)"
    write_ws['A6'] = "메뉴"
    write_ws['A7'] = "리뷰 리스트" # add titles
    for key, value in dic.items():
        tmpchr = ''
        for k in alphabetValue[::-1]:
            tmpchr += chr(k) 

        write_ws[tmpchr+'2'] = (key)
        write_ws[tmpchr+'3'] = value[0]
        write_ws[tmpchr+'4'] = value[1]
        write_ws[tmpchr+'5'] = value[2]
        write_ws[tmpchr+'6'] = value[3]

        for n in range(0, len(value[4])):
            write_ws[tmpchr+str(n+7)] = value[4][n]

        for m in range(0, len(alphabetValue)):
            if(alphabetValue[m] == ord('Z')):
                if(m == len(alphabetValue) - 1):
                    alphabetValue.append(ord('A'))
                alphabetValue[m] = ord('A')
            else:
                alphabetValue[m] += 1 # add info 

    write_wb.save(filename= food+'.xlsx') # save the file

else:
    print("검색 결과를 찾을 수 없습니다.") # if list length is zero, print error message

